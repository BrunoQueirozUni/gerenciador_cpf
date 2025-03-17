import openpyxl  # Biblioteca para manipulação de arquivos Excel
from selenium import webdriver  # Biblioteca para automação de navegadores web
from selenium.webdriver.common.by import By  # Módulo para localizar elementos na página
from time import sleep  # Função para pausar a execução do código

# Abre a planilha de clientes e extrai os dados
planilhaClientes = openpyxl.load_workbook("dados_clientes.xlsx")  # Carrega a planilha Excel
paginaClientes = planilhaClientes["Sheet1"]  # Seleciona a primeira aba da planilha

# Inicializa o driver do navegador (Chrome) e abre a URL especificada
driver = webdriver.Chrome()
driver.get("https://consultcpf-devaprender.netlify.app/")

# Itera sobre as linhas da planilha, começando da segunda linha
for linha in paginaClientes.iter_rows(min_row=2, values_only=True):
    print(linha)  # Imprime a linha atual para depuração

    nome, valor, cpf, vencimento = linha  # Desempacota os valores da linha
    sleep(5)  # Pausa de 5 segundos para garantir o carregamento do site

    # Localiza o campo de pesquisa pelo CPF e insere o valor do CPF
    campoPesquisa = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    sleep(1)  # Pausa para garantir que o campo foi carregado
    campoPesquisa.clear()  # Limpa o campo de pesquisa
    campoPesquisa.send_keys(cpf)  # Insere o CPF no campo de pesquisa
    sleep(1)  # Pausa para garantir que o CPF foi inserido

    # Localiza e clica no botão de consulta
    botaoConsultar = driver.find_element(By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)  # Pausa para garantir que o botão foi carregado
    botaoConsultar.click()  # Clica no botão de consulta
    sleep(4)  # Pausa para aguardar a resposta da consulta

    # Verifica o status da consulta
    campoStatus = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
    
    if campoStatus.text == "em dia":
        # Se o status for "em dia", extrai a data e o método de pagamento
        dataPagamento = driver.find_element(By.XPATH, "//p[@id='paymentDate']")
        metodoPagamento = driver.find_element(By.XPATH, "//p[@id='paymentMethod']")
        
        dataPagamentoData = dataPagamento.text.split()[3]  # Extrai a data de pagamento
        metodoPagamentoMetodo = metodoPagamento.text.split()[3]  # Extrai o método de pagamento
        
        # Abre a planilha de fechamento e adiciona os dados
        planilhaFechamento = openpyxl.load_workbook("dados_clientes_fechamento.xlsx")
        paginaFechamento = planilhaFechamento["Sheet1"]
        paginaFechamento.append([nome, valor, cpf, vencimento, "em dia", dataPagamentoData, metodoPagamentoMetodo])
        
        # Salva a planilha de fechamento
        planilhaFechamento.save("dados_clientes_fechamento.xlsx")
    else:
        # Se o status não for "em dia", adiciona os dados com status "pendente"
        planilhaFechamento = openpyxl.load_workbook("dados_clientes_fechamento.xlsx")
        paginaFechamento = planilhaFechamento["Sheet1"]
        paginaFechamento.append([nome, valor, cpf, vencimento, "pendente"])
        
        # Salva a planilha de fechamento
        planilhaFechamento.save("dados_clientes_fechamento.xlsx")